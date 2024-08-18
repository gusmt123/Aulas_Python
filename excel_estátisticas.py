import numpy as np
import openpyxl

caminho  = 'notas.xlsx'

planilha = openpyxl.load_workbook(caminho)

for nova_aba in planilha.sheetnames:

    planilha_atual = planilha[nova_aba]

    for coluna in planilha_atual.iter_cols:

        media = np.mean(coluna)
        mediana = np.median(coluna)
        maximo = np.max(coluna)
        minimo = np.min(coluna)

        print(media)
        print(mediana)
        print(maximo)
        print(minimo)