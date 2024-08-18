import openpyxl#importa o openpyxl, biblioteca para trabalhar com excel no python
import numpy as np#importa a biblioteca numpy usada para matemática comom a variável np


file_path = '../caixa.xlsx'  # Caminho para a planilha


planilha = openpyxl.load_workbook(file_path)# Carrega o arquivo Excel usandoa biblioteca Openpyxl


for nome_aba in planilha.sheetnames:#Repete o loop para cada planilha
    
    sheet = planilha[nome_aba]#cria a variável sheet que é igual a aba que está atualmente rodando no loop

    print('\n----- ABA:' + nome_aba + '-----')#printa a aba atual

    
    for coluna in sheet.iter_cols(min_col=1, max_col=sheet.max_column):#loopa para cada coluna da aba atual
       
        dados_filtrados = [celula.value for celula in coluna if celula.value is not None 
                            and isinstance(celula.value, (int, float))
                            and celula.value not in [0, 1]] #filtra dados inválidos como strings, valores em branco ou 0 ou 1

        #se o filtro tiver sido feita executa e printa os cálculos
        if dados_filtrados:
            media = np.mean(dados_filtrados)#cálcula a média
            mediana = np.median(dados_filtrados)#cálcula a mediana
            quartil_25 = np.percentile(dados_filtrados, 25)#cálcula o quartil de 25%
            quartil_75 = np.percentile(dados_filtrados, 75)#cálcula o quartil de 75%
            maximo = np.max(dados_filtrados)#mostra o valor máximo da coluna
            minimo = np.min(dados_filtrados)#mostra o valor mínimo da coluna

            # Printa os resultados
            print('\nColuna: ' + str(coluna[0].column_letter))
            print('Média: ' + str(media))
            print('Mediana:' + str(mediana))
            print('Quartil 0.25:' + str(quartil_25))
            print('Quartil 0.75: ' + str(quartil_75))
            print('Máximo: ' + str(maximo))
            print('Mínimo: '+ str(minimo))


planilha.close()# Fecha a planilha após o uso
