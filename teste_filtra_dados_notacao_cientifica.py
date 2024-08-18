import openpyxl
import mysql.connector

# Configurações da conexão
config = {
    'user': 'root',          # Usuário
    'host': 'localhost',     # Host
    'port': '3306',          # Porta
    'database': 'dados',     # Nome do banco
}

# Conectar ao banco de dados MySQL
conexao = mysql.connector.connect(**config)

if conexao.is_connected():
    print('Conectado ao servidor MySQL')

caminho_arquivo = 'notacao_cientifica.xlsx'

# Carregar a planilha Excel
planilha = openpyxl.load_workbook(caminho_arquivo)

for nome_aba in planilha.sheetnames:
    planilha_atual = planilha[nome_aba]
    
    for coluna in planilha_atual.iter_cols(min_col=1, max_col=planilha_atual.max_column):
        for celula in coluna:
            if celula.value is not None and isinstance(celula.value, float) and 'E' in str(celula.value):
                celula.value = float(celula.value)
    
    # Salvar as alterações
    planilha.save(caminho_arquivo)

    cursor = conexao.cursor()

    # Iterar sobre as células da planilha atual
    for linha in planilha_atual.iter_linhas(min_linha=2, max_linha=planilha_atual.max_linha, min_col=1, max_col=planilha_atual.max_column):
        for celula in linha:
            # Verificar se a célula contém um valor numérico
            if isinstance(celula.value, (int, float)):
                # Inserir o valor no banco de dados
                cursor.execute("INSERT INTO numeros (numero) VALUES (%s)", (celula.value,))

    # Commit das alterações
    conexao.commit()

    print(f'Dados da aba "{nome_aba}" inseridos no banco de dados.')

# Fechar conexão com o banco de dados
conexao.close()

 

