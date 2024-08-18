import openpyxl
import mysql.connector

# Configurações da conexão
config = {
    'user': 'root',          # Usuário
    'host': 'localhost',     # Host
    'port': '3306',          # Porta
    'database': 'dados',     # Nome do banco
}

# Conectar ao banco de dados MySQL
conexao = mysql.connector.connect(**config)

if conexao.is_connected():
    print('Conectado ao servidor MySQL')

caminho_arquivo = 'notacao_cientifica.xlsx'

planilha = openpyxl.load_workbook(caminho_arquivo)  # Abre planilha

for nome_aba in planilha.sheetnames:
    planilha_atual = planilha[nome_aba]
    cursor = conexao.cursor()

    # Loop para cada linha
    for linha in planilha_atual.iter_rows(min_row=2, max_row=planilha_atual.max_row, min_col=1, max_col=planilha_atual.max_column):
        for celula in linha:
            # Verificar se a célula contém um valor numérico
            if isinstance(celula.value, (int, float)):
                # Converter para string formatada sem notação científica
                valor_formatado = "{:.9f}".format(celula.value)  # Ajuste a precisão conforme necessário

                # Inserir o valor no banco de dados
                sql = "INSERT INTO numeros (numero) VALUES (%s)"
                cursor.execute(sql, (valor_formatado,))

    conexao.commit()
    print('Dados da aba ' + nome_aba + ' inseridos no banco de dados.')

conexao.close()  # Fecha conexão com o banco de dados
