import openpyxl#importa o openpyxl, biblioteca para trabalhar com excel no python
import numpy as np#importa a biblioteca numpy usada para matemática comom a variável np

caminho_do_arquivo = 'planilha_alunos.xlsx'

planilha = openpyxl.load_workbook(caminho_do_arquivo)

for nome_aba in planilha.sheetnames:
  
    planilha_atual =  planilha[nome_aba]
    
    for coluna in planilha_atual.iter_cols(min_col=1, max_col=planilha_atual.max_column):
        
        
        dados_filtrados = [celula.value for celula in coluna if celula.value is not None and isinstance(celula.value, (int, float))]    
        #para cada celula na copluna que não esteja nulo e seja dos tipos de dados int e float

        if dados_filtrados:

            media = np.mean(dados_filtrados)
            mediana = np.median(dados_filtrados)
            maior = np.max(dados_filtrados)
            menor = np.min(dados_filtrados)

            print('coluna: ' + coluna[0].value)
            print('Média: ' + str(media))
            print('Mediana: ' + str(mediana))
            print('Maior: ' + str(maior))
            print('Menor: ' + str(menor))

        valor_minimo = 2
        valor_maximo = 8
        dados_filtrados_2 = [celula.value for celula in coluna if celula.value is not None and isinstance(celula.value, (int, float)) and valor_minimo <= celula.value <= valor_maximo]

        if dados_filtrados_2:

            media = np.mean(dados_filtrados_2)
            mediana = np.median(dados_filtrados_2)
            maior = np.max(dados_filtrados_2)
            menor = np.min(dados_filtrados_2)

            print('coluna: ' + coluna[0].value)
            print('Média: ' + str(media))
            print('Mediana: ' + str(mediana))
            print('Maior: ' + str(maior))
            print('Menor: ' + str(menor))
